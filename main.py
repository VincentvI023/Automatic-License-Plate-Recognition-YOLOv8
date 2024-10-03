from ultralytics import YOLO
import cv2
from datetime import datetime, timedelta

import util
from sort.sort import *
from util import get_car, read_license_plate
import os
import csv
import openpyxl
import numpy as np
from tqdm import tqdm

plate_folder = '/plates'
results = {}

mot_tracker = Sort()

# load models
coco_model = YOLO('yolov8n.pt')
license_plate_detector = YOLO('license_plate_detector.pt')

# load video
cap = cv2.VideoCapture('./videos/3-10-2024-ochtend.MOV')

# Starttijd invoeren (bijvoorbeeld: de tijd waarop de video is begonnen)
start_time_str = "07:38"  # Starttijd in het formaat "HH:MM"
start_time = datetime.strptime(start_time_str, "%H:%M")  # Omzetten naar datetime object

# Clear all rows except the first one in the CSV file
# csv_file_path = './test.csv'
# with open(csv_file_path, 'r') as file:
#     reader = csv.reader(file)
#     header = next(reader)

# with open(csv_file_path, 'w', newline='') as file:
#     writer = csv.writer(file)
#     writer.writerow(header)

vehicles = [2, 3, 5, 7]
vehicle_info = [
    {"id": 2, "name": "car", "count": 0},
    {"id": 3, "name": "motorcycle", "count": 0},
    {"id": 5, "name": "bus", "count": 0},
    {"id": 7, "name": "truck", "count": 0}
]
unique_vehicle_ids = set()
track_id_to_vehicle_type = {}

# read frames
frame_nmr = -1
ret = True

# Get total number of frames
total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

# Initialize progress bar
progress_bar = tqdm(total=total_frames, desc="Processing frames")

while ret:
    frame_nmr += 1
    ret, frame = cap.read()
    progress_bar.update(1)
    frame_nmr += 1
    ret, frame = cap.read()
    if ret:
        results[frame_nmr] = {}
        # detect vehicles
        detections = coco_model(frame)[0]
        detections_ = []
        for detection in detections.boxes.data.tolist():
            x1, y1, x2, y2, score, class_id = detection
            if int(class_id) in vehicles:
                detections_.append([x1, y1, x2, y2, score])  
                

            # track vehicles
        if len(detections_) > 0: 
            track_ids = mot_tracker.update(np.asarray(detections_))
            
        # Loop over tracked objects and add track_id to unique_vehicle_ids
        for tracked_object in track_ids:
            track_id = int(tracked_object[4])  # track_id is in the 5th position (index 4)
            
            # Detect vehicles using bounding boxes and get corresponding car
            for detection in detections.boxes.data.tolist():
                x1, y1, x2, y2, score, class_id = detection
                
                # Only process if the detected class_id corresponds to a vehicle
                if int(class_id) in vehicles:
                    # Use get_car to associate this detection with a vehicle track_id
                    car_values = get_car([x1, y1, x2, y2, score, class_id], track_ids)
                    _, _, _, _, car_track_id = car_values
                    
                    # If get_car returns a valid track_id (not -1), check if it's unique
                    if car_track_id != -1 and car_track_id not in unique_vehicle_ids:
                        # Add the unique track_id to the set
                        unique_vehicle_ids.add(car_track_id)
                        
                        # Record the vehicle class (car, bus, etc.)
                        track_id_to_vehicle_type[car_track_id] = int(class_id)
                        
                        # Increment the count for the corresponding vehicle class
                        for vehicle in vehicle_info:
                            if vehicle["id"] == int(class_id):
                                vehicle["count"] += 1  # Increment count for this vehicle type
                                # Write the updated vehicle counts to the Excel sheet
                                excel_file_path = './vehicle_counts.xlsx'

                                # Load the workbook and select the active sheet
                                if os.path.exists(excel_file_path):
                                    workbook = openpyxl.load_workbook(excel_file_path)
                                    sheet = workbook.active
                                else:
                                    workbook = openpyxl.Workbook()
                                    sheet = workbook.active
                                    sheet.append(["Vehicle Type", "Count"])

                                # Find the row corresponding to the vehicle type and update the count
                                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
                                    if row[0].value == vehicle["name"]:
                                        row[1].value = vehicle["count"]
                                        break
                                else:
                                    # If the vehicle type is not found, append a new row
                                    sheet.append([vehicle["name"], vehicle["count"]])

                                # Save the workbook
                                workbook.save(excel_file_path)
                                break

        # After processing, print the unique vehicle counts
        print(f"Total unique vehicles: {len(unique_vehicle_ids)}")
        for vehicle in vehicle_info:
            print(f"Total {vehicle['name']}s: {vehicle['count']}")

        # detect license plates
        license_plates = license_plate_detector(frame)[0]
        for license_plate in license_plates.boxes.data.tolist():
            x1, y1, x2, y2, score, class_id = license_plate

            # assign license plate to car
            xcar1, ycar1, xcar2, ycar2, car_id = get_car(license_plate, track_ids)

            if car_id != -1:

                # crop license plate
                license_plate_crop = frame[int(y1):int(y2), int(x1): int(x2), :]
                cv2.imwrite(os.path.join(plate_folder, f'frame_{frame_nmr}_car_{car_id}_crop.jpg'), license_plate_crop)

                # process license plate
                license_plate_crop_gray = cv2.cvtColor(license_plate_crop, cv2.COLOR_BGR2GRAY)
                cv2.imwrite(os.path.join(plate_folder, f'frame_{frame_nmr}_car_{car_id}_gray.jpg'), license_plate_crop_gray)

                license_plate_crop_thresh = cv2.adaptiveThreshold(license_plate_crop_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 13, 2)
                cv2.imwrite(os.path.join(plate_folder, f'frame_{frame_nmr}_car_{car_id}_thresh.jpg'), license_plate_crop_thresh)

                # Additional processing variants
                license_plate_crop_blur = cv2.GaussianBlur(license_plate_crop_gray, (5, 5), 0)
                cv2.imwrite(os.path.join(plate_folder, f'frame_{frame_nmr}_car_{car_id}_blur.jpg'), license_plate_crop_blur)

                _, license_plate_crop_binary = cv2.threshold(license_plate_crop_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                cv2.imwrite(os.path.join(plate_folder, f'frame_{frame_nmr}_car_{car_id}_binary.jpg'), license_plate_crop_binary)

                license_plate_crop_median = cv2.medianBlur(license_plate_crop_gray, 5)
                cv2.imwrite(os.path.join(plate_folder, f'frame_{frame_nmr}_car_{car_id}_median.jpg'), license_plate_crop_median)

                kernel = np.ones((3, 3), np.uint8)
                license_plate_crop_morph = cv2.morphologyEx(license_plate_crop_thresh, cv2.MORPH_CLOSE, kernel)
                cv2.imwrite(os.path.join(plate_folder, f'frame_{frame_nmr}_car_{car_id}_morph.jpg'), license_plate_crop_morph)

                # Perform OCR on different variants of the license plate
                license_plate_text_variants = [
                    read_license_plate(license_plate_crop),
                    read_license_plate(license_plate_crop_gray),
                    read_license_plate(license_plate_crop_thresh),
                    read_license_plate(license_plate_crop_blur),
                    read_license_plate(license_plate_crop_binary),
                    read_license_plate(license_plate_crop_median),
                    read_license_plate(license_plate_crop_morph)
                ]

                # Select the best variant based on the text score
                license_plate_text, license_plate_text_score, car_values = max(license_plate_text_variants, key=lambda x: x[1] if x[0] is not None else 0)
                
                if license_plate_text is not None and license_plate_text_score > 0.55:
                    print("--------------------------------------------------")
                    print(license_plate_text, license_plate_text_score, car_values[0], car_values[1])
                    print("--------------------------------------------------")

                    # Get current time in the video (milliseconds)
                    current_time_ms = cap.get(cv2.CAP_PROP_POS_MSEC)

                    # Convert milliseconds to seconds and create timedelta
                    video_time_delta = timedelta(milliseconds=current_time_ms)

                    # Add video time to the start time
                    detection_time = start_time + video_time_delta
                    
                    # Format detection time as string
                    formatted_detection_time = detection_time.strftime("%H:%M:%S.%f")[:-3]

                    excel_file_path = './car_values_test_day.xlsx'

                    # Load the workbook and select the active sheet
                    if os.path.exists(excel_file_path):
                        workbook = openpyxl.load_workbook(excel_file_path)
                        sheet = workbook.active

                    # Append the new row with car values and actual time
                    sheet.append([license_plate_text, license_plate_text_score, car_values[0], car_values[1], formatted_detection_time])
                    workbook.save(excel_file_path)
                    
                    # Write results to CSV
                    # with open(csv_file_path, 'a', newline='') as file:
                    #     writer = csv.writer(file)
                    #     writer.writerow([frame_nmr, car_id, xcar1, ycar1, xcar2, ycar2, x1, y1, x2, y2, license_plate_text, score, license_plate_text_score, formatted_detection_time])
                else:
                    continue      
print("End of processing")
